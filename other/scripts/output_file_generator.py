import argparse
import logging
import os
import openpyxl.styles
import pandas as pd
import re
import openpyxl


logging.basicConfig(level=logging.INFO,
                    format='%(asctime)s - %(levelname)s - %(message)s')

template_headings = []

def find_markdown_files(directory, excluded_dirs, keyword):
    """
    Traverse the directory to find markdown files that include a specific keyword.

    Args:
        directory (str): The root directory to start the search.
        excluded_dirs (list): List of directories to exclude from the search.
        keyword (str): Keyword to filter markdown files.

    Returns:
        list: List of markdown file paths that match the criteria.
    """
    markdown_files = []
    for root, dirs, files in os.walk(directory):
        dirs[:] = [d for d in dirs if d not in excluded_dirs]
        for file in files:
            if file.endswith(".md") and keyword in file:
                markdown_files.append(os.path.join(root, file))
    return markdown_files


def write_to_excel(df, output_excel):
    """
    Write the DataFrame to an Excel file and format it.

    Args:
        df (pandas.DataFrame): DataFrame to write to Excel.
        output_csv (str): Path to the output Excel file.
    """
    df.to_excel(output_excel, index=False)


def write_to_csv(df, output_csv):
    """
    Write the DataFrame to a CSV file.

    Args:
        df (pandas.DataFrame): DataFrame to write to CSV.
        output_csv (str): Path to the output CSV file.
    """
    df.to_csv(output_csv, index=False)


def extract_headings_and_content(file_path, is_template_file = False):
    """
    Extract headings and their content from a markdown file.

    Args:
        file_path (str): Path to the markdown file.
        is_template_file: Specifies when the template file is initially loaded.
    Returns:
        dict: Dictionary with headings as keys and their content as values.
    """
    name_pattern = re.compile(r"# ([a-zA-Z0-9., +\-\ \/(\)]*)\s*## BB Tag")
    heading_pattern = re.compile(r"## ([a-zA-Z +\-\/\(\)]*)")

    with open(file_path, "r", encoding="utf-8") as file:
        content = file.read()

    content = re.sub(r"<!--.*?-->", "", content, flags=re.DOTALL)
    headings = heading_pattern.findall(content)
    names = name_pattern.findall(content)

    if not names:
        raise ValueError(f"No BB Name found in {file_path}")

    implementation_status = 'implementation exists'
    if "WorkInProgress" in file_path:
        implementation_status = 'suggested'

    heading_contents = {"BB Name": names[0], "Implementation Status":implementation_status}

    for i, heading in enumerate(headings):
        if heading not in template_headings and not is_template_file:
            logging.warning(f"Wrong template in {file_path}, contains unspecified heading: {heading}.")
            continue
        start = content.find(heading) + len(heading)
        end = content.find(headings[i + 1]) if i + \
            1 < len(headings) else len(content)
        heading_content = content[start: end - 3].replace("\n", " ").strip()
        if heading_content.startswith(("-", "+", "=")):
            heading_content = "'" + heading_content
        heading_contents[heading] = heading_content
    if len(heading_contents) < len(template_headings) and not is_template_file:
        template_diff = list(set(template_headings).difference(heading_contents))
        logging.warning(f"Missing heading(s) specified in template in file {file_path}. Missing headings: {template_diff}")
    return heading_contents


def create_df(markdown_files):
    """
    Create a DataFrame from a list of markdown files.

    Args:
        markdown_files (list): List of markdown file paths.

    Returns:
        pandas.DataFrame: DataFrame containing the extracted headings and content.
    """
    file_contents = []

    for file in markdown_files:
        heading_content = extract_headings_and_content(file)
        file_contents.append(heading_content)

    df = pd.DataFrame.from_dict(file_contents)
    return df

def format_excel_file(file, output_excel):
    """
    If an Excel file is generated as an output, it is formatted here to be more readable.
    The first row of text is colored and text wrapping is implemented. In addtion, the column
    width is adjusted.

    Args:
        file (str): Path to generated Excel output file.
        output_excel: (df): Pandas Data Frame from which Excel file is generated, used as a 
        helper file here to provide number of columns to fill etc.

    """
    wb = openpyxl.load_workbook(file)
    ws = wb.active

    empty_border_style = openpyxl.styles.Side(border_style=None)
    no_border = openpyxl.styles.borders.Border(
    left=empty_border_style, 
    right=empty_border_style, 
    top=empty_border_style, 
    bottom=empty_border_style,
    )

    for row in ws:
        for cell in row:
            cell.border = no_border
            cell.alignment = openpyxl.styles.Alignment(wrap_text=True)

    num_rows = output_excel.shape[0]
    num_cols = output_excel.shape[1]
    blue_fill = openpyxl.styles.PatternFill(start_color='68aad4', end_color='68aad4', fill_type='solid')

    i = 1
    while i <= num_rows:
        ws.row_dimensions[i].height = 40
        i = i+1

    i = 1
    while i <= num_cols:
        column_letter = ws.cell(row=1, column=i).column_letter
        ws.column_dimensions[column_letter].width = 40
        i = i+1
    
    
    for col in ws.iter_cols(max_col=num_cols, min_row=1, max_row=1):
        for cell in col:
            cell.fill = blue_fill

    wb.save(file)

def main():
    parser = argparse.ArgumentParser(
        description="Generate output file with specified format from BB template markdown files.")
    parser.add_argument(
        "directory", help="Directory to traverse")
    parser.add_argument(
        "-o", "--output", help="Output export file path, default is 'output'", default="output")

    parser.add_argument("--exclude", nargs="*", default=[".github", ".venv", "other", "UseCases"],
                        help="Directories to exclude, default is '.github', '.venv', 'other', 'UseCases'")
    parser.add_argument("--keyword", default="BB",
                        help="Keyword to filter markdown files, default is 'BB'")
    parser.add_argument("--file_format", default='xlsx', choices=["xlsx", "csv"],
                        help="Format of the output file. Default is 'xlsx'")

    args = parser.parse_args()

    try:
        template_file = os.path.join(args.directory, "other/utils/BB_Template.md")
        global template_headings
        template_headings = [entry for entry in extract_headings_and_content(template_file, is_template_file=True)]
        markdown_files = find_markdown_files(
            args.directory, args.exclude, args.keyword)
        df = create_df(markdown_files)
        output_file_path = args.output + "." + args.file_format

        if args.file_format == 'csv':
            write_to_csv(df, output_file_path)
            logging.info(f"CSV file successfully created at {output_file_path}")
        elif args.file_format == 'xlsx':
            write_to_excel(df, output_file_path)
            format_excel_file(output_file_path, df)
            logging.info(f"Excel file successfully created at {output_file_path}")
    except Exception as e:
        logging.error(f"An error occurred: {e}")


if __name__ == "__main__":
    main()
